"""
rl_environment_test_complete.py

Complete integration test suite for the Finance RL Backend.

This version tests BOTH:
1. the endpoints used by the AGENT_BOT during training/evaluation; and
2. the administrative/non-agent endpoints that are still part of the backend.

The suite uses two authenticated users:
- AGENT_BOT for the RL-facing interface;
- ADMIN for permission-protected maintenance/business endpoints.

Current contract assumptions:
- API base path: /api
- POST  /sandbox/reset accepts an optional deterministic seed
- PATCH /approval/approve
- PATCH /approval/reject
- POST  /invoice/duplicate-check includes invoiceId
- GET   /report/transactions uses normal query parameters (?status=...)
- randomization is part of /sandbox/reset; there is no separate /sandbox/randomize test
- manager approval is no longer a separate endpoint

The script also writes a detailed Excel log under results/logs/test/.
"""

from __future__ import annotations

import json
import os
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import dotenv
import requests
from openpyxl import Workbook, load_workbook

dotenv.load_dotenv()


# ============================================================
# CONFIGURATION
# ============================================================


def _resolve_urls() -> Tuple[str, str]:
    """Return (server_root_url, api_base_url)."""

    configured = os.getenv("BASE_URL", "").strip().rstrip("/")

    if configured:
        if configured.endswith("/api"):
            return configured[:-4], configured

        return configured, f"{configured}/api"

    port = os.getenv("PORT", "5000")
    root = f"http://localhost:{port}"
    return root, f"{root}/api"


SERVER_URL, API_BASE_URL = _resolve_urls()

REQUEST_TIMEOUT = int(os.getenv("TIMEOUT", "30"))
TEST_SEED = int(os.getenv("RANDOM_SEED", "42"))

HEADERS = {
    "Content-Type": "application/json",
}


# ============================================================
# TEST USER
# ============================================================

TEST_USER = {
    "username": os.getenv("TEST_USERNAME", "RL_Test_Agent"),
    "email": os.getenv("EMAIL", "rl_agent_test@sandbox.com"),
    "password": os.getenv("PASSWORD", "Password123!"),
    "role": "AGENT_BOT",
}

# Separate privileged account used only to exercise endpoints that the
# AGENT_BOT is intentionally not allowed to call.
TEST_ADMIN = {
    "username": os.getenv("ADMIN_TEST_USERNAME", "RL_Test_Admin"),
    "email": os.getenv("ADMIN_EMAIL", "rl_admin_test@sandbox.com"),
    "password": os.getenv("ADMIN_PASSWORD", "Password123!"),
    "role": "ADMIN",
}


# ============================================================
# GLOBAL TEST STATE
# ============================================================

TOKEN: Optional[str] = None
ADMIN_TOKEN: Optional[str] = None
EPISODE_ID: Optional[str] = None

INVOICES: List[Dict[str, Any]] = []
SUPPLIERS: List[Dict[str, Any]] = []
ACCOUNTS: List[Dict[str, Any]] = []

INVOICE_ID: Optional[str] = None
DUPLICATE_CHECK_INVOICE: Optional[Dict[str, Any]] = None
APPROVAL_INVOICE_ID: Optional[str] = None
PAYMENT_INVOICE_ID: Optional[str] = None
RETRY_INVOICE_ID: Optional[str] = None

SUPPLIER_ID: Optional[str] = None
ACCOUNT_ID: Optional[str] = None
TO_ACCOUNT_ID: Optional[str] = None

TRANSACTION_ID: Optional[str] = None
CANCEL_TRANSACTION_ID: Optional[str] = None
REFUND_TRANSACTION_ID: Optional[str] = None

CREATED_INVOICE_ID: Optional[str] = None
UPDATE_INVOICE_ID: Optional[str] = None
ARCHIVE_INVOICE_ID: Optional[str] = None
REJECT_INVOICE_ID: Optional[str] = None
BLACKLIST_SUPPLIER_ID: Optional[str] = None


STATS = {
    "total_actions": 0,
    "successful_actions": 0,
    "failed_actions": 0,
    "environment_errors": 0,
    "total_reward": 0.0,
}


# ============================================================
# LOG FILE CONFIGURATION
# ============================================================

LOG_FILE = (
    Path("results")
    / "logs"
    / "test"
    / f"rl_environment_test_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
)


def init_log() -> None:
    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Endpoint Tests"

    sheet.append(
        [
            "Timestamp",
            "Phase",
            "Endpoint",
            "Method",
            "Status Code",
            "HTTP OK",
            "Business Success",
            "Environment Error",
            "Error Type",
            "Reward",
            "Execution Time(ms)",
            "Response",
            "Message",
        ]
    )

    workbook.save(LOG_FILE)


def _safe_json(response: requests.Response) -> Dict[str, Any]:
    try:
        data = response.json()
    except Exception:
        return {"raw_response": response.text}

    return data if isinstance(data, dict) else {"data": data}


def write_log(
    endpoint: str,
    method: str,
    response: requests.Response,
    execution_time: float = 0.0,
    phase: str = "API Execution",
) -> None:
    if not LOG_FILE.exists():
        init_log()

    workbook = load_workbook(LOG_FILE)
    sheet = workbook["Endpoint Tests"]

    data = _safe_json(response)

    sheet.append(
        [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            phase,
            endpoint,
            method,
            response.status_code,
            bool(response.ok),
            data.get("success"),
            bool(
                data.get(
                    "environmentError",
                    data.get("environment_error", False),
                )
            ),
            data.get("errorType", data.get("error_type")),
            data.get("reward"),
            round(execution_time, 2),
            json.dumps(data, indent=2, default=str),
            data.get("message", data.get("reason", "")),
        ]
    )

    workbook.save(LOG_FILE)


# ============================================================
# TERMINAL OUTPUT
# ============================================================


def print_header(title: str) -> None:
    print(f"\n{'=' * 70}")
    print(title)
    print("=" * 70)


def print_result(name: str, success: bool, detail: Optional[str] = None) -> None:
    status = "PASS" if success else "FAIL"
    suffix = f" - {detail}" if detail else ""
    print(f"[{status}] {name}{suffix}")


# ============================================================
# RESPONSE / OBJECT HELPERS
# ============================================================


def _environment_error(data: Dict[str, Any], response: requests.Response) -> bool:
    return bool(
        response.status_code >= 500
        or data.get(
            "environmentError",
            data.get("environment_error", False),
        )
    )


def endpoint_responded(response: Optional[requests.Response]) -> bool:
    """
    True when the backend endpoint itself responded normally.

    A 4xx business response can still prove that the route/method/body reached
    the correct controller. Infrastructure errors (5xx/environmentError) do
    not count as a healthy endpoint response.
    """

    if response is None:
        return False

    data = _safe_json(response)
    return not _environment_error(data, response)


def business_success(response: Optional[requests.Response]) -> bool:
    if response is None:
        return False

    data = _safe_json(response)

    return bool(
        response.ok
        and not _environment_error(data, response)
        and data.get("success") is True
    )


def _object_id(value: Any) -> Optional[str]:
    if value is None:
        return None

    if isinstance(value, str):
        return value

    if isinstance(value, dict):
        candidate = value.get("_id", value.get("id"))
        return str(candidate) if candidate is not None else None

    return str(value)


def _supplier_id_from_invoice(invoice: Dict[str, Any]) -> Optional[str]:
    return _object_id(
        invoice.get(
            "supplier",
            invoice.get("supplierId"),
        )
    )


def _invoice_status(invoice: Dict[str, Any]) -> Optional[str]:
    return invoice.get("status", invoice.get("invoiceStatus"))


def _supplier_is_usable(supplier: Dict[str, Any]) -> bool:
    active = supplier.get(
        "active",
        supplier.get("isActive", True),
    )

    try:
        risk_score = float(supplier.get("riskScore", 0) or 0)
    except (TypeError, ValueError):
        risk_score = 0.0

    return bool(active) and risk_score <= 70


def _account_is_usable(account: Dict[str, Any]) -> bool:
    return not bool(account.get("frozen", False))


def _account_balance(account: Dict[str, Any]) -> float:
    raw = account.get(
        "balance",
        account.get("currentBalance", 0),
    )

    try:
        return float(raw or 0)
    except (TypeError, ValueError):
        return 0.0


def _extract_list(data: Dict[str, Any], *keys: str) -> List[Dict[str, Any]]:
    for key in keys:
        value = data.get(key)

        if isinstance(value, list):
            return [item for item in value if isinstance(item, dict)]

        if isinstance(value, dict):
            for nested_key in keys:
                nested = value.get(nested_key)
                if isinstance(nested, list):
                    return [item for item in nested if isinstance(item, dict)]

    return []


def _extract_transaction_id(data: Dict[str, Any]) -> Optional[str]:
    transaction = data.get("transaction")

    if isinstance(transaction, dict):
        tx_id = _object_id(transaction)
        if tx_id:
            return tx_id

    state = data.get("state")

    if isinstance(state, dict):
        transaction = state.get("transaction")
        if isinstance(transaction, dict):
            tx_id = _object_id(transaction)
            if tx_id:
                return tx_id

        tx_id = state.get("transactionId")
        if tx_id:
            return str(tx_id)

    tx_id = data.get("transactionId")
    if tx_id:
        return str(tx_id)

    return None


def _find_invoice(invoice_id: Optional[str]) -> Optional[Dict[str, Any]]:
    if not invoice_id:
        return None

    for invoice in INVOICES:
        if _object_id(invoice) == str(invoice_id):
            return invoice

    return None


# ============================================================
# REWARD / ACTION STATISTICS
# ============================================================


def update_statistics(response: requests.Response) -> None:
    """Track agent-facing endpoint outcomes while excluding environment errors."""

    data = _safe_json(response)

    if _environment_error(data, response):
        STATS["environment_errors"] += 1
        return

    reward = data.get("reward")

    if isinstance(reward, (int, float)):
        STATS["total_reward"] += float(reward)

    STATS["total_actions"] += 1

    if data.get("success") is True:
        STATS["successful_actions"] += 1
    else:
        STATS["failed_actions"] += 1


# ============================================================
# HTTP REQUEST WRAPPER
# ============================================================


def api_request(
    method: str,
    endpoint: str,
    body: Optional[Dict[str, Any]] = None,
    *,
    params: Optional[Dict[str, Any]] = None,
    auth: bool = False,
    token: Optional[str] = None,
    api: bool = True,
    track_stats: bool = True,
    phase: str = "API Execution",
) -> Optional[requests.Response]:
    """
    Common API caller.

    `endpoint` is relative to `/api` when `api=True`.
    Example:
        api_request("GET", "/invoice", auth=True)
    """

    headers = HEADERS.copy()

    if auth:
        selected_token = token or TOKEN
        if selected_token:
            headers["Authorization"] = f"Bearer {selected_token}"

    base = API_BASE_URL if api else SERVER_URL
    url = f"{base}{endpoint}"

    display_endpoint = f"/api{endpoint}" if api else endpoint
    start_time = time.time()

    try:
        response = requests.request(
            method=method.upper(),
            url=url,
            headers=headers,
            json=body if method.upper() not in {"GET", "HEAD"} else None,
            params=params,
            timeout=REQUEST_TIMEOUT,
        )

        execution_time = (time.time() - start_time) * 1000

        if track_stats:
            update_statistics(response)

        write_log(
            display_endpoint,
            method.upper(),
            response,
            execution_time,
            phase=phase,
        )

        return response

    except requests.exceptions.RequestException as exc:
        STATS["environment_errors"] += 1
        print(f"Request error for {display_endpoint}: {exc}")
        return None


# ============================================================
# 1. HEALTH CHECK
# ============================================================


def test_health() -> bool:
    response = api_request(
        "GET",
        "/",
        api=False,
        track_stats=False,
        phase="Health Check",
    )

    success = bool(
        response is not None
        and response.status_code == 200
        and _safe_json(response).get("success") is True
    )

    print_result("Server health", success)
    return success


# ============================================================
# 2. AUTHENTICATION
# ============================================================


def test_register() -> bool:
    response = api_request(
        "POST",
        "/auth/register",
        TEST_USER,
        track_stats=False,
        phase="Authentication",
    )

    if response is None:
        print_result("Register test user", False)
        return False

    data = _safe_json(response)

    # Re-running the integration test should not fail simply because the
    # dedicated test account already exists.
    already_exists = (
        response.status_code in {400, 409}
        and "exist" in str(data.get("message", "")).lower()
    )

    success = data.get("success") is True or already_exists
    print_result("Register test user", success)
    return success


def test_login() -> bool:
    global TOKEN

    response = api_request(
        "POST",
        "/auth/login",
        {
            "email": TEST_USER["email"],
            "password": TEST_USER["password"],
        },
        track_stats=False,
        phase="Authentication",
    )

    if response is None:
        print_result("Login", False)
        return False

    data = _safe_json(response)

    token = data.get("token")

    nested_data = data.get("data")
    if not token and isinstance(nested_data, dict):
        token = nested_data.get("token")

    if token:
        TOKEN = str(token)

    success = bool(data.get("success") is True and TOKEN)
    print_result("Login", success)
    return success


def test_profile() -> bool:
    response = api_request(
        "GET",
        "/auth/me",
        auth=True,
        track_stats=False,
        phase="Authentication",
    )

    success = business_success(response)
    print_result("Get authenticated profile", success)
    return success


def test_authentication() -> bool:
    register_ok = test_register()
    login_ok = test_login()

    if not login_ok:
        return False

    profile_ok = test_profile()
    return register_ok and login_ok and profile_ok


def test_register_admin() -> bool:
    response = api_request(
        "POST",
        "/auth/register",
        TEST_ADMIN,
        track_stats=False,
        phase="Admin Authentication",
    )

    if response is None:
        print_result("Register admin test user", False)
        return False

    data = _safe_json(response)
    already_exists = (
        response.status_code in {400, 409}
        and "exist" in str(data.get("message", "")).lower()
    )

    success = data.get("success") is True or already_exists
    print_result("Register admin test user", success)
    return success


def test_login_admin() -> bool:
    global ADMIN_TOKEN

    response = api_request(
        "POST",
        "/auth/login",
        {
            "email": TEST_ADMIN["email"],
            "password": TEST_ADMIN["password"],
        },
        track_stats=False,
        phase="Admin Authentication",
    )

    if response is None:
        print_result("Login admin", False)
        return False

    data = _safe_json(response)
    token_value = data.get("token")

    nested = data.get("data")
    if not token_value and isinstance(nested, dict):
        token_value = nested.get("token")

    if token_value:
        ADMIN_TOKEN = str(token_value)

    success = bool(data.get("success") is True and ADMIN_TOKEN)
    print_result("Login admin", success)
    return success


def test_admin_profile() -> bool:
    response = api_request(
        "GET",
        "/auth/me",
        auth=True,
        token=ADMIN_TOKEN,
        track_stats=False,
        phase="Admin Authentication",
    )

    if response is None:
        print_result("Get admin profile", False)
        return False

    data = _safe_json(response)
    user = data.get("user") if isinstance(data.get("user"), dict) else {}
    success = bool(
        business_success(response) and (not user or user.get("role") == "ADMIN")
    )
    print_result("Get admin profile", success)
    return success


def test_admin_authentication() -> bool:
    register_ok = test_register_admin()
    login_ok = test_login_admin()

    if not login_ok:
        return False

    profile_ok = test_admin_profile()
    return register_ok and login_ok and profile_ok


# ============================================================
# 3. EPISODE START
# ============================================================


def test_start_episode() -> bool:
    global EPISODE_ID

    payload = {
        "agentType": "RL",
        "algorithm": "PPO",
        "goal": "Backend endpoint integration test",
        "initialState": {},
        "experimentName": "backend_endpoint_test",
        "phase": "TEST",
        "seed": TEST_SEED,
    }

    response = api_request(
        "POST",
        "/episode/start",
        payload,
        auth=True,
        track_stats=False,
        phase="Episode Logging",
    )

    if response is None:
        print_result("Start episode", False)
        return False

    data = _safe_json(response)

    EPISODE_ID = data.get("episodeId")

    if not EPISODE_ID:
        episode = data.get("episode")
        if isinstance(episode, dict):
            EPISODE_ID = _object_id(episode)

    success = bool(data.get("success") is True and EPISODE_ID)
    print_result(
        "Start episode", success, f"episode={EPISODE_ID}" if EPISODE_ID else None
    )
    return success


# ============================================================
# 4. SANDBOX
# ============================================================


def test_reset_environment() -> bool:
    payload: Dict[str, Any] = {
        "seed": TEST_SEED,
    }

    if EPISODE_ID:
        payload["episodeId"] = EPISODE_ID

    response = api_request(
        "POST",
        "/sandbox/reset",
        payload,
        auth=True,
        track_stats=False,
        phase="Sandbox",
    )

    if response is None:
        print_result("Reset environment", False)
        return False

    data = _safe_json(response)

    seed_ok = True
    if data.get("seed") is not None:
        try:
            seed_ok = int(data["seed"]) == TEST_SEED
        except (TypeError, ValueError):
            seed_ok = False

    success = bool(
        response.ok
        and data.get("success") is True
        and not _environment_error(data, response)
        and seed_ok
    )

    print_result("Reset environment with seed", success, f"seed={TEST_SEED}")
    return success


def test_get_state() -> bool:
    response = api_request(
        "GET",
        "/sandbox/state",
        auth=True,
        track_stats=False,
        phase="Sandbox",
    )

    if response is None:
        print_result("Get sandbox state", False)
        return False

    data = _safe_json(response)
    observation = data.get("observation", data.get("state"))

    success = bool(
        data.get("success") is True
        and isinstance(observation, dict)
        and not _environment_error(data, response)
    )

    print_result("Get sandbox state", success)
    return success


def test_sandbox_workflow() -> bool:
    return test_reset_environment() and test_get_state()


# ============================================================
# 5. SUPPLIER ENDPOINTS
# ============================================================


def test_get_suppliers() -> bool:
    global SUPPLIERS, SUPPLIER_ID, BLACKLIST_SUPPLIER_ID

    response = api_request(
        "GET",
        "/supplier",
        auth=True,
        track_stats=False,
        phase="Supplier",
    )

    if response is None:
        print_result("Get suppliers", False)
        return False

    data = _safe_json(response)
    SUPPLIERS = _extract_list(data, "suppliers", "data")

    preferred = next(
        (supplier for supplier in SUPPLIERS if _supplier_is_usable(supplier)),
        SUPPLIERS[0] if SUPPLIERS else None,
    )

    SUPPLIER_ID = _object_id(preferred) if preferred else None

    # Keep a different supplier for the destructive blacklist endpoint when
    # possible so normal agent-facing tests are unaffected.
    blacklist_candidate = next(
        (
            supplier
            for supplier in reversed(SUPPLIERS)
            if _object_id(supplier) != SUPPLIER_ID
        ),
        preferred,
    )
    BLACKLIST_SUPPLIER_ID = (
        _object_id(blacklist_candidate) if blacklist_candidate else None
    )

    success = bool(data.get("success") is True and SUPPLIERS)
    print_result("Get suppliers", success, f"count={len(SUPPLIERS)}")
    return success


def test_validate_supplier() -> bool:
    if not SUPPLIER_ID:
        print_result("Validate supplier", False, "No supplier fixture")
        return False

    response = api_request(
        "POST",
        "/supplier/validate",
        {"supplierId": SUPPLIER_ID},
        auth=True,
        phase="Supplier",
    )

    # A randomized supplier can be inactive/high-risk. The route is considered
    # healthy as long as this is a business outcome rather than an environment
    # failure. We preferentially selected a low-risk active supplier above.
    success = endpoint_responded(response)
    print_result("Validate supplier", success)
    return success


def test_supplier_workflow() -> bool:
    return test_get_suppliers() and test_validate_supplier()


# ============================================================
# 6. INVOICE ENDPOINTS
# ============================================================


def _select_invoice_fixtures() -> None:
    global INVOICE_ID
    global DUPLICATE_CHECK_INVOICE
    global APPROVAL_INVOICE_ID
    global PAYMENT_INVOICE_ID
    global RETRY_INVOICE_ID

    if not INVOICES:
        return

    # First general-purpose invoice.
    INVOICE_ID = _object_id(INVOICES[0])

    # Prefer a non-rejected invoice for duplicate checking.
    DUPLICATE_CHECK_INVOICE = next(
        (
            invoice
            for invoice in INVOICES
            if _invoice_status(invoice) != "REJECTED"
            and _supplier_id_from_invoice(invoice)
            and invoice.get("amount") is not None
            and invoice.get("dueDate")
        ),
        None,
    )

    pending_candidates = [
        invoice
        for invoice in INVOICES
        if _invoice_status(invoice) == "PENDING_APPROVAL"
    ]

    approved_candidates = [
        invoice for invoice in INVOICES if _invoice_status(invoice) == "APPROVED"
    ]

    # Prefer candidates that are not explicitly flagged as duplicates.
    pending_candidates.sort(key=lambda inv: bool(inv.get("duplicateFlag", False)))
    approved_candidates.sort(key=lambda inv: bool(inv.get("duplicateFlag", False)))

    if pending_candidates:
        APPROVAL_INVOICE_ID = _object_id(pending_candidates[0])

    if approved_candidates:
        PAYMENT_INVOICE_ID = _object_id(approved_candidates[0])

    RETRY_INVOICE_ID = PAYMENT_INVOICE_ID or APPROVAL_INVOICE_ID or INVOICE_ID


def test_get_invoices() -> bool:
    global INVOICES

    response = api_request(
        "GET",
        "/invoice",
        auth=True,
        phase="Invoice",
    )

    if response is None:
        print_result("Get invoices", False)
        return False

    data = _safe_json(response)
    INVOICES = _extract_list(data, "invoices", "data")
    _select_invoice_fixtures()

    success = bool(data.get("success") is True and INVOICES)
    print_result("Get invoices", success, f"count={len(INVOICES)}")
    return success


def test_get_invoice_by_id() -> bool:
    if not INVOICE_ID:
        print_result("Get invoice by id", False, "No invoice fixture")
        return False

    response = api_request(
        "GET",
        f"/invoice/{INVOICE_ID}",
        auth=True,
        track_stats=False,
        phase="Invoice",
    )

    success = business_success(response)
    print_result("Get invoice by id", success)
    return success


def test_duplicate_check() -> bool:
    invoice = DUPLICATE_CHECK_INVOICE

    if not invoice:
        print_result("Duplicate check", False, "No suitable invoice fixture")
        return False

    invoice_id = _object_id(invoice)
    supplier_id = _supplier_id_from_invoice(invoice)
    amount = invoice.get("amount")
    due_date = invoice.get("dueDate")

    payload = {
        "invoiceId": invoice_id,
        "supplierId": supplier_id,
        "amount": amount,
        "dueDate": due_date,
    }

    response = api_request(
        "POST",
        "/invoice/duplicate-check",
        payload,
        auth=True,
        phase="Invoice",
    )

    if response is None:
        print_result("Duplicate check", False)
        return False

    data = _safe_json(response)

    success = bool(
        response.ok
        and data.get("success") is True
        and "duplicate" in data
        and not _environment_error(data, response)
    )

    print_result(
        "Duplicate check with invoiceId",
        success,
        f"duplicate={data.get('duplicate')}" if response is not None else None,
    )
    return success


def test_invoice_workflow() -> bool:
    return test_get_invoices() and test_get_invoice_by_id() and test_duplicate_check()


# ============================================================
# 7. ACCOUNT / BUDGET ENDPOINTS
# ============================================================


def test_get_accounts() -> bool:
    global ACCOUNTS, ACCOUNT_ID, TO_ACCOUNT_ID

    response = api_request(
        "GET",
        "/account",
        auth=True,
        track_stats=False,
        phase="Account",
    )

    if response is None:
        print_result("Get accounts", False)
        return False

    data = _safe_json(response)
    ACCOUNTS = _extract_list(data, "accounts", "data")

    usable_accounts = [account for account in ACCOUNTS if _account_is_usable(account)]

    if usable_accounts:
        usable_accounts.sort(
            key=_account_balance,
            reverse=True,
        )
        ACCOUNT_ID = _object_id(usable_accounts[0])

        if len(usable_accounts) > 1:
            TO_ACCOUNT_ID = _object_id(usable_accounts[1])
    elif ACCOUNTS:
        ACCOUNT_ID = _object_id(ACCOUNTS[0])
        if len(ACCOUNTS) > 1:
            TO_ACCOUNT_ID = _object_id(ACCOUNTS[1])

    success = bool(data.get("success") is True and ACCOUNTS)
    print_result("Get accounts", success, f"count={len(ACCOUNTS)}")
    return success


def _budget_fixture() -> Tuple[str, Optional[float]]:
    preferred = (
        _find_invoice(PAYMENT_INVOICE_ID)
        or _find_invoice(APPROVAL_INVOICE_ID)
        or (INVOICES[0] if INVOICES else None)
    )

    if not preferred:
        return "SOFTWARE", None

    department = preferred.get(
        "category",
        preferred.get("department", "SOFTWARE"),
    )

    amount = preferred.get("amount")
    return str(department or "SOFTWARE"), amount


def test_budget_check() -> bool:
    department, amount = _budget_fixture()

    payload: Dict[str, Any] = {
        "department": department,
    }

    if amount is not None:
        payload["amount"] = amount

    response = api_request(
        "POST",
        "/account/budget/check",
        payload,
        auth=True,
        phase="Account",
    )

    if response is None:
        print_result("Budget check", False)
        return False

    data = _safe_json(response)

    # In the corrected backend, finding that a department is missing or that
    # an amount is outside budget is a successful validation result rather
    # than an API/action failure.
    success = bool(
        response.ok
        and data.get("success") is True
        and not _environment_error(data, response)
        and any(
            key in data
            for key in (
                "budget",
                "found",
                "eligible",
                "withinBudget",
                "remainingBudget",
            )
        )
    )

    print_result("Budget check", success, f"department={department}")
    return success


def test_cash_position() -> bool:
    response = api_request(
        "GET",
        "/account/cash-position",
        auth=True,
        track_stats=False,
        phase="Account",
    )

    if response is None:
        print_result("Cash position", False)
        return False

    data = _safe_json(response)

    success = bool(
        response.ok
        and data.get("success") is True
        and not _environment_error(data, response)
        and ("cashPosition" in data or "accounts" in data)
    )

    print_result("Cash position", success)
    return success


def test_account_workflow() -> bool:
    return test_get_accounts() and test_budget_check() and test_cash_position()


# ============================================================
# 8. APPROVAL ENDPOINT
# ============================================================


def test_approve_invoice() -> bool:
    global PAYMENT_INVOICE_ID

    if not APPROVAL_INVOICE_ID:
        # A random reset should normally contain pending invoices, but if it
        # does not, the route cannot be exercised meaningfully.
        print_result("Approve invoice (PATCH)", False, "No pending invoice fixture")
        return False

    response = api_request(
        "PATCH",
        "/approval/approve",
        {"invoiceId": APPROVAL_INVOICE_ID},
        auth=True,
        phase="Approval",
    )

    if response is None:
        print_result("Approve invoice (PATCH)", False)
        return False

    data = _safe_json(response)

    success = business_success(response)

    if success:
        PAYMENT_INVOICE_ID = APPROVAL_INVOICE_ID

    print_result(
        "Approve invoice (PATCH /approval/approve)",
        success,
        data.get("message"),
    )
    return success


# ============================================================
# 9. REPORTING ENDPOINTS
# ============================================================


def _transactions_from_response(data: Dict[str, Any]) -> List[Dict[str, Any]]:
    return _extract_list(data, "transactions", "data")


def test_transactions_report() -> bool:
    global CANCEL_TRANSACTION_ID

    response = api_request(
        "GET",
        "/report/transactions",
        params={"status": "PENDING"},
        auth=True,
        track_stats=False,
        phase="Reporting",
    )

    if response is None:
        print_result("Get transactions with ?status=PENDING", False)
        return False

    data = _safe_json(response)
    transactions = _transactions_from_response(data)

    pending = next(
        (
            transaction
            for transaction in transactions
            if transaction.get("status") == "PENDING"
        ),
        transactions[0] if transactions else None,
    )

    if pending:
        CANCEL_TRANSACTION_ID = _object_id(pending) or pending.get("transactionId")
        if CANCEL_TRANSACTION_ID is not None:
            CANCEL_TRANSACTION_ID = str(CANCEL_TRANSACTION_ID)

    success = bool(
        response.ok
        and data.get("success") is True
        and not _environment_error(data, response)
    )

    print_result(
        "Get transactions with standard query parameter",
        success,
        f"count={len(transactions)}",
    )
    return success


def test_generate_report() -> bool:
    end_date = datetime.now(timezone.utc)
    start_date = end_date - timedelta(days=365)

    response = api_request(
        "POST",
        "/report/generate-report",
        {
            "type": "transaction_summary",
            "startDate": start_date.isoformat(),
            "endDate": end_date.isoformat(),
        },
        auth=True,
        phase="Reporting",
    )

    if response is None:
        print_result("Generate transaction report", False)
        return False

    data = _safe_json(response)

    success = bool(
        response.ok
        and data.get("success") is True
        and not _environment_error(data, response)
        and ("report" in data or "data" in data)
    )

    print_result("Generate transaction_summary report", success)
    return success


def test_reporting_workflow() -> bool:
    return test_transactions_report() and test_generate_report()


# ============================================================
# 10. PAYMENT ENDPOINTS
# ============================================================


def _refresh_invoices() -> bool:
    global INVOICES

    response = api_request(
        "GET",
        "/invoice",
        auth=True,
        track_stats=False,
        phase="Fixture Refresh",
    )

    if not business_success(response):
        return False

    data = _safe_json(response)
    INVOICES = _extract_list(data, "invoices", "data")
    return bool(INVOICES)


def _candidate_payment_invoices() -> Iterable[Dict[str, Any]]:
    approved = [
        invoice for invoice in INVOICES if _invoice_status(invoice) == "APPROVED"
    ]

    # Prefer obviously eligible invoices first.
    def score(invoice: Dict[str, Any]) -> Tuple[int, float]:
        duplicate = int(bool(invoice.get("duplicateFlag", False)))

        supplier = invoice.get("supplier")
        supplier_bad = 0
        if isinstance(supplier, dict):
            supplier_bad = int(not _supplier_is_usable(supplier))

        try:
            amount = float(invoice.get("amount", 0) or 0)
        except (TypeError, ValueError):
            amount = 0.0

        return duplicate + supplier_bad, amount

    return sorted(approved, key=score)


def _choose_payment_invoice() -> Optional[str]:
    if not INVOICES:
        _refresh_invoices()

    for invoice in _candidate_payment_invoices():
        invoice_id = _object_id(invoice)

        if not invoice_id:
            continue

        if invoice.get("duplicateFlag") is True:
            continue

        supplier = invoice.get("supplier")
        if isinstance(supplier, dict) and not _supplier_is_usable(supplier):
            continue

        if ACCOUNT_ID:
            account = next(
                (a for a in ACCOUNTS if _object_id(a) == ACCOUNT_ID),
                None,
            )

            if account:
                try:
                    amount = float(invoice.get("amount", 0) or 0)
                except (TypeError, ValueError):
                    amount = 0.0

                if _account_balance(account) < amount:
                    continue

        return invoice_id

    # Fallback: at least exercise the endpoint with an approved invoice.
    approved = next(iter(_candidate_payment_invoices()), None)
    return _object_id(approved) if approved else PAYMENT_INVOICE_ID


def test_pay_invoice() -> bool:
    global PAYMENT_INVOICE_ID, TRANSACTION_ID

    _refresh_invoices()
    PAYMENT_INVOICE_ID = _choose_payment_invoice()

    if not PAYMENT_INVOICE_ID or not ACCOUNT_ID:
        print_result("Pay invoice", False, "Missing approved invoice/account fixture")
        return False

    response = api_request(
        "POST",
        "/payment/pay",
        {
            "invoiceId": PAYMENT_INVOICE_ID,
            "accountId": ACCOUNT_ID,
        },
        auth=True,
        phase="Payment",
    )

    if response is None:
        print_result("Pay invoice", False)
        return False

    data = _safe_json(response)
    TRANSACTION_ID = _extract_transaction_id(data)

    # For this integration suite, a deterministic business conflict still
    # confirms that the route and request contract are correct. A full business
    # success is reported in the detail so it is easy to distinguish.
    success = endpoint_responded(response)

    print_result(
        "Pay invoice",
        success,
        (
            "business success"
            if data.get("success") is True
            else f"business outcome={data.get('errorType', data.get('message'))}"
        ),
    )
    return success


def test_cancel_payment() -> bool:
    transaction_id = CANCEL_TRANSACTION_ID or TRANSACTION_ID

    if not transaction_id:
        print_result("Cancel payment", False, "No transaction fixture")
        return False

    response = api_request(
        "POST",
        "/payment/cancel-payment",
        {"transactionId": transaction_id},
        auth=True,
        phase="Payment",
    )

    success = endpoint_responded(response)
    print_result("Cancel payment", success)
    return success


def test_retry_payment() -> bool:
    invoice_id = RETRY_INVOICE_ID or PAYMENT_INVOICE_ID or INVOICE_ID

    if not invoice_id:
        print_result("Retry payment", False, "No invoice fixture")
        return False

    response = api_request(
        "POST",
        "/payment/retry-payment",
        {"invoiceId": invoice_id},
        auth=True,
        phase="Payment",
    )

    success = endpoint_responded(response)
    print_result("Retry payment", success)
    return success


def test_payment_workflow() -> bool:
    pay = test_pay_invoice()
    cancel = test_cancel_payment()
    retry = test_retry_payment()

    return pay and cancel and retry


# ============================================================
# 11. ADMINISTRATIVE / NON-AGENT ENDPOINTS
# ============================================================


def _admin_request(
    method: str,
    endpoint: str,
    body: Optional[Dict[str, Any]] = None,
    *,
    params: Optional[Dict[str, Any]] = None,
    track_stats: bool = False,
    phase: str = "Administrative",
) -> Optional[requests.Response]:
    if not ADMIN_TOKEN:
        return None

    return api_request(
        method,
        endpoint,
        body,
        params=params,
        auth=True,
        token=ADMIN_TOKEN,
        track_stats=track_stats,
        phase=phase,
    )


def _extract_invoice_from_response(data: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    invoice = data.get("invoice")
    if isinstance(invoice, dict):
        return invoice

    state = data.get("state")
    if isinstance(state, dict) and isinstance(state.get("invoice"), dict):
        return state["invoice"]

    nested = data.get("data")
    if isinstance(nested, dict):
        invoice = nested.get("invoice")
        if isinstance(invoice, dict):
            return invoice

    return None


def _create_admin_invoice(
    purpose: str,
    *,
    amount: float = 1.0,
    category: str = "SOFTWARE",
) -> Optional[str]:
    if not SUPPLIER_ID:
        return None

    payload = {
        "supplierId": SUPPLIER_ID,
        "amount": amount,
        "category": category,
        "description": f"Backend integration test - {purpose}",
        "dueDate": "2026-12-31",
        "paymentMethod": "BANK",
        "priority": "LOW",
    }

    response = _admin_request(
        "POST",
        "/invoice",
        payload,
        phase="Administrative Invoice",
    )

    if response is None:
        return None

    data = _safe_json(response)
    if data.get("success") is not True:
        return None

    invoice = _extract_invoice_from_response(data)
    if invoice:
        return _object_id(invoice)

    invoice_id = data.get("invoiceId")
    return str(invoice_id) if invoice_id else None


def _find_affordable_budget_department(amount: float = 1.0) -> Optional[str]:
    # Use known invoice categories first, then the full model enum.
    candidates: List[str] = []

    for invoice in INVOICES:
        category = invoice.get("category")
        if category and category not in candidates:
            candidates.append(str(category))

    for category in ("SOFTWARE", "HARDWARE", "TRAVEL", "SERVICES"):
        if category not in candidates:
            candidates.append(category)

    for department in candidates:
        response = api_request(
            "POST",
            "/account/budget/check",
            {"department": department, "amount": amount},
            auth=True,
            track_stats=False,
            phase="Administrative Fixture",
        )

        if response is None:
            continue

        data = _safe_json(response)
        if (
            response.ok
            and data.get("success") is True
            and data.get("found", True) is not False
            and data.get("eligible", True) is not False
            and data.get("withinBudget", True) is not False
        ):
            return department

    return None


def test_create_invoice_admin() -> bool:
    global CREATED_INVOICE_ID

    category = _find_affordable_budget_department(1.0) or "SOFTWARE"
    CREATED_INVOICE_ID = _create_admin_invoice(
        "create endpoint",
        amount=1.0,
        category=category,
    )

    success = CREATED_INVOICE_ID is not None
    print_result(
        "Create invoice (ADMIN)",
        success,
        f"invoice={CREATED_INVOICE_ID}" if CREATED_INVOICE_ID else None,
    )
    return success


def test_update_invoice_status_admin() -> bool:
    global UPDATE_INVOICE_ID

    category = _find_affordable_budget_department(1.0) or "SOFTWARE"
    UPDATE_INVOICE_ID = _create_admin_invoice(
        "update status endpoint",
        amount=1.0,
        category=category,
    )

    if not UPDATE_INVOICE_ID:
        print_result("Update invoice status (ADMIN)", False, "Could not create fixture")
        return False

    response = _admin_request(
        "PATCH",
        f"/invoice/{UPDATE_INVOICE_ID}/status",
        {"status": "APPROVED"},
        phase="Administrative Invoice",
    )

    success = business_success(response)
    print_result("Update invoice status (ADMIN)", success)
    return success


def test_archive_invoice_admin() -> bool:
    global ARCHIVE_INVOICE_ID

    category = _find_affordable_budget_department(1.0) or "SOFTWARE"
    ARCHIVE_INVOICE_ID = _create_admin_invoice(
        "archive endpoint",
        amount=1.0,
        category=category,
    )

    if not ARCHIVE_INVOICE_ID:
        print_result("Archive invoice (ADMIN)", False, "Could not create fixture")
        return False

    response = _admin_request(
        "POST",
        "/invoice/archive",
        {"invoiceId": ARCHIVE_INVOICE_ID},
        phase="Administrative Invoice",
    )

    success = business_success(response)
    print_result("Archive invoice (ADMIN)", success)
    return success


def test_reject_invoice_admin() -> bool:
    global REJECT_INVOICE_ID

    category = _find_affordable_budget_department(1.0) or "SOFTWARE"
    REJECT_INVOICE_ID = _create_admin_invoice(
        "reject endpoint",
        amount=1.0,
        category=category,
    )

    if not REJECT_INVOICE_ID:
        print_result("Reject invoice (ADMIN)", False, "Could not create fixture")
        return False

    response = _admin_request(
        "PATCH",
        "/approval/reject",
        {
            "invoiceId": REJECT_INVOICE_ID,
            "reason": "Backend integration test rejection",
        },
        phase="Administrative Approval",
    )

    success = business_success(response)
    print_result("Reject invoice (PATCH, ADMIN)", success)
    return success


def test_blacklist_supplier_admin() -> bool:
    if not BLACKLIST_SUPPLIER_ID:
        print_result("Blacklist supplier (ADMIN)", False, "No supplier fixture")
        return False

    response = _admin_request(
        "POST",
        "/supplier/blacklist",
        {
            "supplierId": BLACKLIST_SUPPLIER_ID,
            "reason": "Backend integration test blacklist",
        },
        phase="Administrative Supplier",
    )

    success = business_success(response)
    print_result("Blacklist supplier (ADMIN)", success)
    return success


def test_account_transfer_admin() -> bool:
    if not ACCOUNT_ID or not TO_ACCOUNT_ID:
        print_result("Account transfer (ADMIN)", False, "Need two account fixtures")
        return False

    source = next(
        (account for account in ACCOUNTS if _object_id(account) == ACCOUNT_ID),
        None,
    )

    if source is None:
        print_result("Account transfer (ADMIN)", False, "Source account not found")
        return False

    balance = _account_balance(source)
    try:
        transfer_limit = float(source.get("dailyTransferLimit", balance) or balance)
    except (TypeError, ValueError):
        transfer_limit = balance

    amount = min(1.0, balance, transfer_limit)
    if amount <= 0:
        print_result("Account transfer (ADMIN)", False, "No transferable balance")
        return False

    response = _admin_request(
        "POST",
        "/account/transfer",
        {
            "fromAccount": ACCOUNT_ID,
            "toAccount": TO_ACCOUNT_ID,
            "amount": amount,
        },
        phase="Administrative Account",
    )

    success = business_success(response)
    print_result("Account transfer (ADMIN)", success)
    return success


def _create_refundable_payment_fixture() -> Optional[str]:
    if not ACCOUNT_ID or not SUPPLIER_ID:
        return None

    department = _find_affordable_budget_department(1.0)
    if not department:
        return None

    invoice_id = _create_admin_invoice(
        "refund endpoint",
        amount=1.0,
        category=department,
    )
    if not invoice_id:
        return None

    approve_response = _admin_request(
        "PATCH",
        "/approval/approve",
        {"invoiceId": invoice_id},
        phase="Administrative Refund Fixture",
    )
    if not business_success(approve_response):
        return None

    pay_response = api_request(
        "POST",
        "/payment/pay",
        {
            "invoiceId": invoice_id,
            "accountId": ACCOUNT_ID,
        },
        auth=True,
        track_stats=False,
        phase="Administrative Refund Fixture",
    )
    if not business_success(pay_response):
        return None

    return _extract_transaction_id(_safe_json(pay_response))


def test_refund_payment_admin() -> bool:
    global REFUND_TRANSACTION_ID

    REFUND_TRANSACTION_ID = _create_refundable_payment_fixture()

    if not REFUND_TRANSACTION_ID:
        print_result(
            "Refund payment (ADMIN)",
            False,
            "Could not create a successful payment fixture",
        )
        return False

    response = _admin_request(
        "POST",
        "/payment/refund",
        {"transactionId": REFUND_TRANSACTION_ID},
        phase="Administrative Payment",
    )

    success = business_success(response)
    print_result("Refund payment (ADMIN)", success)
    return success


def test_audit_log_admin() -> bool:
    response = _admin_request(
        "GET",
        "/report/audit-log",
        phase="Administrative Reporting",
    )

    if response is None:
        print_result("Get audit log (ADMIN)", False)
        return False

    data = _safe_json(response)
    logs = _extract_list(data, "logs", "auditLogs", "data")

    # An empty audit list is still a valid endpoint response, although this
    # suite normally creates several action logs before reaching this test.
    success = bool(
        response.ok
        and data.get("success") is True
        and not _environment_error(data, response)
        and (
            isinstance(data.get("logs"), list)
            or isinstance(data.get("auditLogs"), list)
            or isinstance(data.get("data"), (list, dict))
            or logs
        )
    )

    print_result("Get audit log (ADMIN)", success, f"count={len(logs)}")
    return success


def test_administrative_workflow() -> bool:
    # Order matters: the destructive blacklist runs last so it cannot affect
    # the invoice/payment fixtures that use the preferred supplier.
    results = [
        test_create_invoice_admin(),
        test_update_invoice_status_admin(),
        test_archive_invoice_admin(),
        test_reject_invoice_admin(),
        test_account_transfer_admin(),
        test_refund_payment_admin(),
        test_audit_log_admin(),
        test_blacklist_supplier_admin(),
    ]
    return all(results)


# ============================================================
# 12. EPISODE LOGGING / REWARD ENDPOINTS
# ============================================================


def test_record_episode_step() -> bool:
    if not EPISODE_ID:
        print_result("Record episode step", False, "No episode id")
        return False

    payload = {
        "step": 1,
        "action": "GET_INVOICES",
        "endpoint": "/api/invoice",
        "reward": 0,
        "success": True,
        "stateBefore": {},
        "stateAfter": {
            "invoiceCount": len(INVOICES),
        },
        "message": "Integration-test episode step.",
    }

    response = api_request(
        "POST",
        f"/episode/{EPISODE_ID}/step",
        payload,
        auth=True,
        track_stats=False,
        phase="Episode Logging",
    )

    success = business_success(response)
    print_result("Record episode step", success)
    return success


def test_end_episode() -> bool:
    if not EPISODE_ID:
        print_result("End episode", False, "No episode id")
        return False

    response = api_request(
        "POST",
        f"/episode/{EPISODE_ID}/end",
        {
            "finalState": {
                "endpointTestCompleted": True,
            },
            "completed": True,
            "terminatedReason": "GOAL_REACHED",
        },
        auth=True,
        track_stats=False,
        phase="Episode Logging",
    )

    success = business_success(response)
    print_result("End episode", success)
    return success


def test_get_episode() -> bool:
    if not EPISODE_ID:
        print_result("Get episode", False, "No episode id")
        return False

    response = api_request(
        "GET",
        f"/episode/{EPISODE_ID}",
        auth=True,
        track_stats=False,
        phase="Episode Logging",
    )

    success = business_success(response)
    print_result("Get episode", success)
    return success


def test_get_episodes() -> bool:
    response = api_request(
        "GET",
        "/episode",
        params={
            "experimentName": "backend_endpoint_test",
            "phase": "TEST",
            "algorithm": "PPO",
        },
        auth=True,
        track_stats=False,
        phase="Episode Logging",
    )

    if response is None:
        print_result("Get episodes", False)
        return False

    data = _safe_json(response)
    success = bool(
        response.ok
        and data.get("success") is True
        and not _environment_error(data, response)
    )

    print_result("Get episodes with filters", success)
    return success


def test_sandbox_reward() -> bool:
    if not EPISODE_ID:
        print_result("Get sandbox reward", False, "No episode id")
        return False

    response = api_request(
        "GET",
        f"/sandbox/reward/{EPISODE_ID}",
        auth=True,
        track_stats=False,
        phase="Sandbox",
    )

    if response is None:
        print_result("Get sandbox reward", False)
        return False

    data = _safe_json(response)

    success = bool(
        response.ok
        and data.get("success") is True
        and "reward" in data
        and not _environment_error(data, response)
    )

    print_result("Get sandbox reward for episode", success)
    return success


def test_episode_workflow() -> bool:
    return all(
        [
            test_record_episode_step(),
            test_end_episode(),
            test_get_episode(),
            test_get_episodes(),
            test_sandbox_reward(),
        ]
    )


# ============================================================
# SUMMARY
# ============================================================


def print_statistics(results: Dict[str, bool]) -> None:
    print_header("TEST SUMMARY")

    for name, result in results.items():
        print_result(name, result)

    print("\nTracked backend action statistics")
    print("-" * 70)
    print(f"Total tracked actions : {STATS['total_actions']}")
    print(f"Successful actions    : {STATS['successful_actions']}")
    print(f"Failed actions        : {STATS['failed_actions']}")
    print(f"Environment errors    : {STATS['environment_errors']}")
    print(f"Total backend reward  : {STATS['total_reward']:.2f}")
    print(f"Excel log             : {LOG_FILE}")


# ============================================================
# RUN ALL TESTS
# ============================================================


def run_all_tests() -> bool:
    init_log()

    results: Dict[str, bool] = {}

    print_header("FINANCE RL BACKEND - UPDATED ENDPOINT INTEGRATION TEST")
    print(f"Server URL : {SERVER_URL}")
    print(f"API URL    : {API_BASE_URL}")
    print(f"Test seed  : {TEST_SEED}")

    results["Health"] = test_health()

    if not results["Health"]:
        print_statistics(results)
        return False

    results["Authentication"] = test_authentication()
    results["Admin Authentication"] = test_admin_authentication()

    if not results["Authentication"] or not results["Admin Authentication"]:
        print_statistics(results)
        return False

    results["Episode Start"] = test_start_episode()
    results["Sandbox"] = test_sandbox_workflow()

    # Load randomized fixtures after reset.
    results["Supplier"] = test_supplier_workflow()
    results["Invoice"] = test_invoice_workflow()
    results["Account"] = test_account_workflow()

    # Approval can turn a pending invoice into a payment candidate.
    results["Approval"] = test_approve_invoice()

    # Reporting finds a PENDING transaction for cancel-payment when available.
    results["Reporting"] = test_reporting_workflow()
    results["Payment"] = test_payment_workflow()

    # Exercise the endpoints that are intentionally outside the AGENT_BOT
    # action interface using the ADMIN test account.
    results["Administrative Endpoints"] = test_administrative_workflow()

    results["Episode Logging"] = test_episode_workflow()

    print_statistics(results)

    return all(results.values())


if __name__ == "__main__":
    success = run_all_tests()
    raise SystemExit(0 if success else 1)
